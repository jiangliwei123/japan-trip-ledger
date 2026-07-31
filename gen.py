# -*- coding: utf-8 -*-
"""
日本旅行记账生成器
读取 data.json，生成：
  1. trip_account.xlsx  - Excel 账目（明细 / 汇总 / AA结算）
  2. index.html         - 手机端交互网页（数据内嵌，可离线打开）
每次成员新增消费，更新 data.json 后重跑本脚本即可同步两文件。
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, "data.json"), encoding="utf-8"))

META = DATA["meta"]
MEMBERS = META["members"]
A, B = MEMBERS[0], MEMBERS[1]
TX = DATA["transactions"]
RATE = float(META.get("currency_rate", 21.0))


def compute():
    """返回各项汇总数据。"""
    a_paid = b_paid = 0.0          # 实际支出
    a_cons = b_cons = 0.0          # 消费（AA 分摊后）
    cat = {}                       # 类别总支出（按金额）
    for t in TX:
        amt = float(t["amount_cny"])
        payer = t["payer"]
        shared = bool(t.get("shared"))
        if payer == A:
            a_paid += amt
        else:
            b_paid += amt
        if shared:
            a_cons += amt / 2
            b_cons += amt / 2
        else:
            if payer == A:
                a_cons += amt
            else:
                b_cons += amt
        cat[t.get("category", "其他")] = cat.get(t.get("category", "其他"), 0.0) + amt
    a_net = a_paid - a_cons
    b_net = b_paid - b_cons
    return dict(a_paid=a_paid, b_paid=b_paid, a_cons=a_cons, b_cons=b_cons,
                a_net=a_net, b_net=b_net, cat=cat,
                total=a_paid + b_paid)


AGG = compute()


# ---------------------------------------------------------------------------
# 1) Excel
# ---------------------------------------------------------------------------
def build_excel(path):
    wb = Workbook()

    # 主题色
    RED = "E60012"
    INK = "1B2A4A"
    PAPER = "F7F5F2"
    GOLD = "C8A35B"
    thin = Side(style="thin", color="D9D2C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor=INK)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color=INK, size=15)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def money(v):
        return round(v, 2)

    # ---- Sheet 1: 消费明细 ----
    ws = wb.active
    ws.title = "消费明细"
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{META['trip']} · 消费明细"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A2"] = f"更新时间：{META.get('updated_at','')}   |   汇率：{META.get('rate_note','')}"
    ws["A2"].font = Font(italic=True, color="8A8A8A", size=9)
    ws.merge_cells("A2:H2")

    headers = ["日期", "付款人", "消费项目", "类别", "金额(元)",
               "是否AA", f"{A}实际支出", f"{B}实际支出"]
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border

    r = hr + 1
    for t in TX:
        amt = float(t["amount_cny"])
        shared = bool(t.get("shared"))
        payer = t["payer"]
        a_p = amt if payer == A else 0.0
        b_p = amt if payer == B else 0.0
        row = [t["date"], payer, t["item"], t.get("category", "其他"),
               money(amt), "AA分摊" if shared else "单人",
               money(a_p), money(b_p)]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = border
            cell.alignment = left if c == 3 else center
            if c in (5, 7, 8):
                cell.alignment = right
                cell.number_format = "¥#,##0.00"
            if c == 2:
                cell.font = Font(bold=True, color=RED if payer == A else INK)
            if c == 6 and shared:
                cell.font = Font(bold=True, color="1B7A3D")
        if r % 2 == 0:
            for c in range(1, 9):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=PAPER)
        r += 1

    # 合计行
    tot_a = AGG["a_paid"]
    tot_b = AGG["b_paid"]
    trow = ["", "", "合计", "", money(AGG["total"]), "", money(tot_a), money(tot_b)]
    for c, v in enumerate(trow, 1):
        cell = ws.cell(r, c, v)
        cell.border = border
        cell.font = Font(bold=True, color=INK)
        cell.fill = PatternFill("solid", fgColor=GOLD)
        cell.alignment = right if c in (5, 7, 8) else (left if c == 3 else center)
        if c in (5, 7, 8):
            cell.number_format = "¥#,##0.00"
    widths = [12, 9, 26, 9, 13, 9, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ---- Sheet 2: 汇总 ----
    ws2 = wb.create_sheet("汇总")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = f"{META['trip']} · 成员汇总"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:D1")

    ws2.append([])
    hdr = ["成员", "实际支出(元)", "消费(元)", "净结(应收+/应付-)"]
    for c, h in enumerate(hdr, 1):
        cell = ws2.cell(3, c, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border
    rows = [
        [A, money(AGG["a_paid"]), money(AGG["a_cons"]), money(AGG["a_net"])],
        [B, money(AGG["b_paid"]), money(AGG["b_cons"]), money(AGG["b_net"])],
    ]
    rr = 4
    for row in rows:
        for c, v in enumerate(row, 1):
            cell = ws2.cell(rr, c, v)
            cell.border = border
            cell.alignment = right if c > 1 else center
            if c > 1:
                cell.number_format = "¥#,##0.00"
            if c == 1:
                cell.font = Font(bold=True, color=RED if row[0] == A else INK)
            if c == 4 and isinstance(v, (int, float)):
                cell.font = Font(bold=True, color="1B7A3D" if v >= 0 else RED)
        rr += 1
    # 总计
    ws2.cell(rr, 1, "总计").font = Font(bold=True)
    ws2.cell(rr, 2, money(AGG["total"])).number_format = "¥#,##0.00"
    ws2.cell(rr, 2).font = Font(bold=True)
    ws2.cell(rr, 2).alignment = right
    ws2.cell(rr, 3, "").border = border
    ws2.cell(rr, 1).border = border
    ws2.cell(rr, 2).border = border

    # 类别构成
    ws2["A" + str(rr + 3)] = "类别支出构成"
    ws2["A" + str(rr + 3)].font = Font(bold=True, color=INK, size=12)
    cr = rr + 4
    ws2.cell(cr, 1, "类别").fill = head_fill
    ws2.cell(cr, 1).font = head_font
    ws2.cell(cr, 2, "金额(元)").fill = head_fill
    ws2.cell(cr, 2).font = head_font
    ws2.cell(cr, 3, "占比").fill = head_fill
    ws2.cell(cr, 3).font = head_font
    ws2.cell(cr, 1).border = border
    ws2.cell(cr, 2).border = border
    ws2.cell(cr, 3).border = border
    ws2.cell(cr, 1).alignment = center
    ws2.cell(cr, 2).alignment = center
    ws2.cell(cr, 3).alignment = center
    cr += 1
    for cat, amt in sorted(AGG["cat"].items(), key=lambda x: -x[1]):
        ws2.cell(cr, 1, cat).border = border
        ws2.cell(cr, 1).alignment = center
        ws2.cell(cr, 2, money(amt)).number_format = "¥#,##0.00"
        ws2.cell(cr, 2).border = border
        ws2.cell(cr, 2).alignment = right
        ws2.cell(cr, 3, amt / AGG["total"]).number_format = "0.0%"
        ws2.cell(cr, 3).border = border
        ws2.cell(cr, 3).alignment = right
        cr += 1
    for i, w in enumerate([14, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet 3: AA结算 ----
    ws3 = wb.create_sheet("AA结算")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = f"{META['trip']} · AA制结算"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:D1")

    a_net = AGG["a_net"]
    if abs(a_net) < 0.005:
        msg = "✅ 已平账，双方无需结算。"
        payer, receiver, amount = "", "", 0.0
    elif a_net > 0:
        msg = f"{B} 需向 {A} 支付，以结清 AA 差额。"
        payer, receiver, amount = B, A, a_net
    else:
        msg = f"{A} 需向 {B} 支付，以结清 AA 差额。"
        payer, receiver, amount = A, B, -a_net

    ws3["A3"] = msg
    ws3["A3"].font = Font(bold=True, color=INK, size=12)
    ws3.merge_cells("A3:D3")

    settle = [
        ("付款人", payer),
        ("收款人", receiver),
        ("结算金额(元)", money(amount) if amount else "—"),
        ("结算金额(日元)", f"¥{money(amount*RATE):,.0f}" if amount else "—"),
    ]
    sr = 5
    for k, v in settle:
        ws3.cell(sr, 1, k).font = Font(bold=True, color="FFFFFF")
        ws3.cell(sr, 1).fill = PatternFill("solid", fgColor=INK)
        ws3.cell(sr, 1).border = border
        ws3.cell(sr, 1).alignment = center
        ws3.cell(sr, 2, v).alignment = left
        ws3.cell(sr, 2).border = border
        ws3.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=4)
        sr += 1
    ws3.column_dimensions["A"].width = 18
    for col in "BCD":
        ws3.column_dimensions[col].width = 14

    wb.save(path)
    print("Excel 已生成:", path)


# ---------------------------------------------------------------------------
# 2) 网页
# ---------------------------------------------------------------------------
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
<title>__TRIP__ · 旅行记账</title>
<style>
  :root{
    --red:#E60012; --ink:#1B2A4A; --paper:#F7F5F2; --gold:#C8A35B;
    --green:#1B7A3D; --muted:#8A8A8A; --card:#FFFFFF; --line:#ECE6DC;
    --shadow:0 6px 20px rgba(27,42,74,.08);
  }
  *{box-sizing:border-box;-webkit-tap-highlight-color:transparent;}
  body{
    margin:0;background:var(--paper);color:var(--ink);
    font-family:-apple-system,BlinkMacSystemFont,"PingFang SC","Segoe UI",Roboto,"Helvetica Neue",sans-serif;
    -webkit-font-smoothing:antialiased;
  }
  .wrap{max-width:480px;margin:0 auto;padding:0 14px 90px;}

  /* 顶部 */
  header{
    position:sticky;top:0;z-index:20;background:linear-gradient(135deg,var(--ink),#2c3e60);
    color:#fff;padding:16px 14px 14px;box-shadow:var(--shadow);
  }
  .head-row{display:flex;align-items:center;justify-content:space-between;max-width:480px;margin:0 auto;}
  .title{font-size:18px;font-weight:700;letter-spacing:.5px;}
  .title small{display:block;font-size:11px;font-weight:400;opacity:.7;margin-top:2px;}
  .toggle{display:flex;background:rgba(255,255,255,.15);border-radius:20px;padding:3px;}
  .toggle button{
    border:0;background:transparent;color:#fff;font-size:13px;font-weight:600;
    padding:6px 14px;border-radius:17px;cursor:pointer;transition:.2s;
  }
  .toggle button.on{background:#fff;color:var(--ink);}

  /* 总支出卡片 */
  .hero{margin:16px 0;background:var(--card);border-radius:18px;padding:18px;box-shadow:var(--shadow);}
  .hero .lbl{font-size:12px;color:var(--muted);}
  .hero .big{font-size:34px;font-weight:800;margin:4px 0 2px;}
  .hero .sub{font-size:12px;color:var(--muted);}

  /* 成员对比 */
  .members{display:flex;gap:12px;margin:14px 0;}
  .mcard{flex:1;background:var(--card);border-radius:16px;padding:14px;box-shadow:var(--shadow);position:relative;overflow:hidden;}
  .mcard::before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;}
  .mcard.a::before{background:var(--red);}
  .mcard.b::before{background:var(--ink);}
  .mname{font-size:14px;font-weight:700;}
  .mname.a{color:var(--red);}
  .mname.b{color:var(--ink);}
  .mrow{display:flex;justify-content:space-between;font-size:12px;margin-top:10px;color:var(--muted);}
  .mrow b{color:var(--ink);font-size:15px;font-weight:700;}
  .mrow .paid b{color:var(--red);}

  /* 区块标题 */
  .sec-title{font-size:14px;font-weight:700;margin:22px 4px 10px;display:flex;align-items:center;gap:6px;}
  .sec-title::before{content:"";width:4px;height:14px;background:var(--gold);border-radius:2px;}

  /* 类别圆环 */
  .cat{background:var(--card);border-radius:16px;padding:16px;box-shadow:var(--shadow);display:flex;align-items:center;gap:16px;}
  .donut{width:120px;height:120px;border-radius:50%;flex:none;position:relative;}
  .donut::after{content:"";position:absolute;inset:26px;background:var(--card);border-radius:50%;}
  .donut .ctr{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;z-index:2;}
  .donut .ctr small{font-size:10px;color:var(--muted);}
  .donut .ctr b{font-size:15px;}
  .legend{flex:1;font-size:12px;}
  .legend div{display:flex;align-items:center;justify-content:space-between;margin:5px 0;}
  .legend .dot{width:9px;height:9px;border-radius:50%;display:inline-block;margin-right:6px;}
  .legend .nm{display:flex;align-items:center;color:var(--ink);}
  .legend .vl{color:var(--muted);}

  /* 明细 */
  .tx{background:var(--card);border-radius:14px;box-shadow:var(--shadow);overflow:hidden;margin-bottom:12px;}
  .tx .day{padding:10px 14px;background:var(--paper);font-size:12px;font-weight:600;color:var(--ink);border-bottom:1px solid var(--line);}
  .tx .item{display:flex;align-items:center;padding:12px 14px;border-bottom:1px solid var(--line);}
  .tx .item:last-child{border-bottom:0;}
  .tx .ic{width:38px;height:38px;border-radius:11px;background:var(--paper);display:flex;align-items:center;justify-content:center;font-size:18px;flex:none;margin-right:11px;}
  .tx .mid{flex:1;min-width:0;}
  .tx .nm{font-size:14px;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
  .tx .meta{font-size:11px;color:var(--muted);margin-top:2px;}
  .tag{font-size:10px;padding:1px 7px;border-radius:9px;margin-left:6px;font-weight:600;}
  .tag.aa{background:#e6f4ea;color:var(--green);}
  .tag.solo{background:#fdeaea;color:var(--red);}
  .tx .amt{text-align:right;flex:none;margin-left:8px;}
  .tx .amt b{font-size:15px;}
  .tx .amt .who{font-size:10px;color:var(--muted);}

  /* 成员筛选 */
  .filter{display:flex;gap:8px;margin:0 4px 12px;}
  .filter button{flex:1;padding:9px 0;border:1px solid var(--line);background:#fff;border-radius:11px;font-size:13px;font-weight:600;color:var(--muted);cursor:pointer;transition:.15s;}
  .filter button.on{background:var(--ink);color:#fff;border-color:var(--ink);}
  .filter button.mA.on{background:var(--red);border-color:var(--red);}
  .filter button.mB.on{background:var(--ink);border-color:var(--ink);}
  .txcap{display:flex;justify-content:space-between;align-items:center;background:#fff7ec;border:1px solid var(--gold);border-radius:12px;padding:10px 14px;margin-bottom:12px;font-size:13px;color:var(--muted);}
  .txcap b{color:var(--ink);font-size:15px;font-weight:700;}

  /* 结算 */
  .settle{background:linear-gradient(135deg,#fff,#fff7ec);border:1px solid var(--gold);border-radius:16px;padding:16px;box-shadow:var(--shadow);}
  .settle .st{font-size:12px;color:var(--muted);}
  .settle .big{font-size:22px;font-weight:800;margin:6px 0;}
  .settle .arrow{display:flex;align-items:center;justify-content:center;gap:10px;margin-top:8px;font-size:14px;}
  .settle .pill{padding:6px 14px;border-radius:20px;font-weight:700;color:#fff;}
  .settle .pill.from{background:var(--red);}
  .settle .pill.to{background:var(--ink);}
  .settle .ok{color:var(--green);font-size:16px;font-weight:700;text-align:center;padding:8px;}

  /* 添加按钮 + 表单 */
  .fab{position:fixed;left:50%;transform:translateX(-50%);bottom:18px;z-index:30;
    background:var(--red);color:#fff;border:0;border-radius:26px;padding:13px 26px;font-size:15px;font-weight:700;
    box-shadow:0 8px 22px rgba(230,0,18,.35);cursor:pointer;}
  .mask{position:fixed;inset:0;background:rgba(27,42,74,.45);z-index:40;display:none;align-items:flex-end;}
  .mask.show{display:flex;}
  .sheet{background:var(--paper);width:100%;max-width:480px;margin:0 auto;border-radius:20px 20px 0 0;padding:20px 18px 26px;animation:up .25s ease;}
  @keyframes up{from{transform:translateY(100%)}to{transform:translateY(0)}}
  .sheet h3{margin:0 0 14px;font-size:16px;}
  .field{margin-bottom:12px;}
  .field label{font-size:12px;color:var(--muted);display:block;margin-bottom:5px;}
  .field input,.field select{width:100%;padding:11px 12px;border:1px solid var(--line);border-radius:10px;font-size:14px;background:#fff;color:var(--ink);}
  .seg{display:flex;gap:8px;}
  .seg button{flex:1;padding:10px;border:1px solid var(--line);background:#fff;border-radius:10px;font-size:13px;font-weight:600;color:var(--muted);}
  .seg button.on{border-color:var(--red);color:var(--red);background:#fff0f0;}
  .sheet .acts{display:flex;gap:10px;margin-top:6px;}
  .sheet .acts button{flex:1;padding:13px;border:0;border-radius:12px;font-size:15px;font-weight:700;cursor:pointer;}
  .btn-add{background:var(--red);color:#fff;}
  .btn-cancel{background:#fff;color:var(--ink);border:1px solid var(--line)!important;}
  .note{font-size:11px;color:var(--muted);margin-top:10px;line-height:1.5;}
  .empty{text-align:center;color:var(--muted);font-size:13px;padding:30px 0;}
</style>
</head>
<body>
<header>
  <div class="head-row">
    <div class="title">__TRIP__<small>双人旅行记账 · AA 结算</small></div>
    <div class="toggle">
      <button id="btnCny" class="on" onclick="setCur('cny')">元</button>
      <button id="btnJpy" onclick="setCur('jpy')">日元</button>
    </div>
  </div>
</header>

<div class="wrap">
  <div class="hero">
    <div class="lbl">总支出</div>
    <div class="big" id="totalAmt">—</div>
    <div class="sub" id="totalSub">—</div>
  </div>

  <div class="members">
    <div class="mcard a">
      <div class="mname a" id="nameA">__A__</div>
      <div class="mrow paid"><span>实际支出</span><b id="aPaid">—</b></div>
      <div class="mrow"><span>消费</span><b id="aCons">—</b></div>
    </div>
    <div class="mcard b">
      <div class="mname b" id="nameB">__B__</div>
      <div class="mrow paid"><span>实际支出</span><b id="bPaid">—</b></div>
      <div class="mrow"><span>消费</span><b id="bCons">—</b></div>
    </div>
  </div>

  <div class="sec-title">支出构成</div>
  <div class="cat">
    <div class="donut" id="donut"><div class="ctr"><small>总支出</small><b id="donutTotal">—</b></div></div>
    <div class="legend" id="legend"></div>
  </div>

  <div class="sec-title">消费明细</div>
  <div class="filter" id="memFilter">
    <button class="on" data-v="all" onclick="setFilter(this)">全部</button>
    <button class="mA" data-v="__A__" onclick="setFilter(this)">__A__</button>
    <button class="mB" data-v="__B__" onclick="setFilter(this)">__B__</button>
  </div>
  <div id="txList"></div>

  <div class="sec-title">AA 制结算</div>
  <div class="settle" id="settle"></div>

  <div class="note" id="rateNote"></div>
</div>

<button class="fab" onclick="openSheet()">＋ 记一笔</button>
<div class="mask" id="mask" onclick="if(event.target===this)closeSheet()">
  <div class="sheet">
    <h3>记录一笔消费</h3>
    <div class="field">
      <label>付款人</label>
      <div class="seg" id="payerSeg">
        <button type="button" data-v="__A__" class="on" onclick="seg(this,'payer')">__A__</button>
        <button type="button" data-v="__B__" onclick="seg(this,'payer')">__B__</button>
      </div>
    </div>
    <div class="field"><label>消费项目</label><input id="fItem" placeholder="如：酒店 / 晚餐 / 门票"></div>
    <div class="field"><label>类别</label>
      <select id="fCat">
        <option>交通</option><option>住宿</option><option>餐饮</option><option>门票</option>
        <option>购物</option><option>签证</option><option>通讯</option><option>其他</option>
      </select>
    </div>
    <div class="field"><label>金额（元）</label><input id="fAmt" type="number" inputmode="decimal" placeholder="0.00"></div>
    <div class="field">
      <label>分摊方式</label>
      <div class="seg" id="shareSeg">
        <button type="button" data-v="aa" class="on" onclick="seg(this,'share')">AA 平分</button>
        <button type="button" data-v="solo" onclick="seg(this,'share')">单人承担</button>
      </div>
    </div>
    <div class="field"><label>日期</label><input id="fDate" type="date"></div>
    <div class="acts">
      <button class="btn-cancel" onclick="closeSheet()">取消</button>
      <button class="btn-add" onclick="addTx()">保存</button>
    </div>
    <div class="note">提示：本表单便于临时记录并保存在本机。正式账目以「姜」告知记账助手为准，助手会同步更新 Excel 与网页。点「重置官方账目」可清除本机临时记录。</div>
    <div class="acts" style="margin-top:8px;">
      <button class="btn-cancel" style="flex:1" onclick="resetLocal()">重置官方账目</button>
    </div>
  </div>
</div>

<script>
const DATA = __DATA_JSON__;
const A = DATA.meta.members[0], B = DATA.meta.members[1];
let RATE = parseFloat(DATA.meta.currency_rate) || 21;
let CUR = 'cny';
let FILTER = 'all';
function setFilter(el){
  document.querySelectorAll('#memFilter button').forEach(b=>b.classList.remove('on'));
  el.classList.add('on');
  FILTER = el.dataset.v;
  render();
}

// 颜色板（类别）
const PAL = ["#E60012","#1B2A4A","#C8A35B","#1B7A3D","#E8743B","#7B61FF","#0FA3B1","#D6336C"];
const ICON = {交通:"🚄",住宿:"🏨",餐饮:"🍣",门票:"🎫",购物:"🛍️",签证:"📋",通讯:"📱",其他:"💴"};

function fmt(v){
  if(CUR==='cny') return '¥'+v.toLocaleString('zh-CN',{minimumFractionDigits:2,maximumFractionDigits:2});
  return '¥'+(v*RATE).toLocaleString('zh-CN',{maximumFractionDigits:0});
}
function calc(tx){
  let aPaid=0,bPaid=0,aCons=0,bCons=0,cat={};
  tx.forEach(t=>{
    const amt=+t.amount_cny, p=t.payer, sh=!!t.shared;
    if(p===A)aPaid+=amt; else bPaid+=amt;
    if(sh){aCons+=amt/2;bCons+=amt/2;} else { if(p===A)aCons+=amt; else bCons+=amt; }
    cat[t.category||'其他']=(cat[t.category||'其他']||0)+amt;
  });
  return {aPaid,bPaid,aCons,bCons,cat,total:aPaid+bPaid,aNet:aPaid-aCons,bNet:bPaid-bCons};
}

function render(){
  const s = calc(localTx());
  document.getElementById('nameA').textContent=A;
  document.getElementById('nameB').textContent=B;
  document.getElementById('totalAmt').textContent=fmt(s.total);
  document.getElementById('totalSub').textContent='约 '+fmt(s.total).replace('¥',CUR==='cny'?'¥':'¥')+'（折合日元 ¥'+(s.total*RATE).toLocaleString('zh-CN',{maximumFractionDigits:0})+'）';
  document.getElementById('aPaid').textContent=fmt(s.aPaid);
  document.getElementById('bPaid').textContent=fmt(s.bPaid);
  document.getElementById('aCons').textContent=fmt(s.aCons);
  document.getElementById('bCons').textContent=fmt(s.bCons);

  // 圆环
  const cats=Object.entries(s.cat).sort((a,b)=>b[1]-a[1]);
  let acc=0, stops=[];
  cats.forEach((c,i)=>{ const pct=c[1]/s.total*360; stops.push(PAL[i%PAL.length]+' '+acc+'deg '+(acc+pct)+'deg'); acc+=pct; });
  const donut=document.getElementById('donut');
  donut.style.background = cats.length? ('conic-gradient('+stops.join(',')+')') : '#ECE6DC';
  document.getElementById('donutTotal').textContent=fmt(s.total);
  const lg=document.getElementById('legend'); lg.innerHTML='';
  cats.forEach((c,i)=>{
    const d=document.createElement('div');
    d.innerHTML='<span class="nm"><span class="dot" style="background:'+PAL[i%PAL.length]+'"></span>'+c[0]+'</span><span class="vl">'+fmt(c[1])+' · '+(c[1]/s.total*100).toFixed(0)+'%</span>';
    lg.appendChild(d);
  });

  // 明细（成员筛选）
  const list=document.getElementById('txList'); list.innerHTML='';
  let rows=[...localTx()];
  if(FILTER!=='all') rows=rows.filter(t=>t.payer===FILTER);
  const sorted=rows.slice().sort((a,b)=>b.date.localeCompare(a.date)||b.id-a.id);
  const fTotal=rows.reduce((s,t)=>s+(+t.amount_cny),0);
  if(!sorted.length){
    list.innerHTML='<div class="empty">'+(FILTER==='all'?'还没有消费记录，点下方按钮记一笔～':'该成员暂无消费记录')+'</div>';
  }
  if(FILTER!=='all' && sorted.length){
    const cap=document.createElement('div'); cap.className='txcap';
    cap.innerHTML='<span>'+FILTER+' 共 '+rows.length+' 笔</span><b>'+fmt(fTotal)+'</b>';
    list.appendChild(cap);
  }
  let curDay='';
  sorted.forEach(t=>{
    if(t.date!==curDay){
      curDay=t.date;
      const dd=document.createElement('div'); dd.className='tx';
      dd.innerHTML='<div class="day">📅 '+t.date+'</div>';
      list.appendChild(dd);
    }
    const who = t.payer===A?A:B;
    const tag = t.shared?'<span class="tag aa">AA</span>':'<span class="tag solo">单人</span>';
    const row=document.createElement('div'); row.className='tx';
    row.innerHTML='<div class="item">'+
      '<div class="ic">'+(ICON[t.category]||'💴')+'</div>'+
      '<div class="mid"><div class="nm">'+t.item+tag+'</div>'+
      '<div class="meta">'+t.category+' · '+who+'支付</div></div>'+
      '<div class="amt"><b>'+fmt(+t.amount_cny)+'</b><div class="who">'+(t.shared?(A+'/'+B+' 各'+(t.amount_cny/2).toFixed(2)):who+'单人')+'</div></div>'+
      '</div>';
    list.appendChild(row);
  });

  // 结算
  const st=document.getElementById('settle');
  if(Math.abs(s.aNet)<0.005){
    st.innerHTML='<div class="st">当前净额</div><div class="ok">✅ 已平账，双方无需结算</div>';
  }else if(s.aNet>0){
    st.innerHTML='<div class="st">'+B+' 多消费，需补偿 '+A+'</div>'+
      '<div class="big">'+fmt(s.aNet)+'</div>'+
      '<div class="arrow"><span class="pill from">'+B+' 支付</span>→<span class="pill to">'+A+' 收款</span></div>';
  }else{
    st.innerHTML='<div class="st">'+A+' 多消费，需补偿 '+B+'</div>'+
      '<div class="big">'+fmt(-s.aNet)+'</div>'+
      '<div class="arrow"><span class="pill from">'+A+' 支付</span>→<span class="pill to">'+B+' 收款</span></div>';
  }

  document.getElementById('rateNote').textContent='汇率说明：'+DATA.meta.rate_note+'　当前 1元 = '+RATE+'日元';
}

function setCur(c){
  CUR=c;
  document.getElementById('btnCny').classList.toggle('on',c==='cny');
  document.getElementById('btnJpy').classList.toggle('on',c==='jpy');
  render();
}

// ---- 本地临时记录 ----
const KEY='jp_trip_tx_'+DATA.meta.trip;
function localTx(){
  try{const l=JSON.parse(localStorage.getItem(KEY)); if(Array.isArray(l)) return l;}catch(e){}
  return JSON.parse(JSON.stringify(DATA.transactions));
}
function saveLocal(arr){localStorage.setItem(KEY,JSON.stringify(arr));}
function resetLocal(){localStorage.removeItem(KEY);render();}

function openSheet(){document.getElementById('mask').classList.add('show');document.getElementById('fDate').value=new Date().toISOString().slice(0,10);}
function closeSheet(){document.getElementById('mask').classList.remove('show');}
let picker={payer:A,share:'aa'};
function seg(el,kind){
  const seg=el.parentElement;
  [...seg.children].forEach(b=>b.classList.remove('on'));
  el.classList.add('on');
  picker[kind]= el.dataset.v;
}
function addTx(){
  const item=document.getElementById('fItem').value.trim();
  const amt=parseFloat(document.getElementById('fAmt').value);
  const date=document.getElementById('fDate').value;
  const cat=document.getElementById('fCat').value;
  if(!item||!(amt>0)||!date){alert('请填写项目、金额与日期');return;}
  const arr=localTx();
  const id=Math.max(0,...arr.map(t=>t.id))+1;
  arr.push({id,date,payer:picker.payer,item,amount_cny:amt,shared:picker.share==='aa',category:cat});
  saveLocal(arr); closeSheet();
  document.getElementById('fItem').value='';document.getElementById('fAmt').value='';
  render();
}

render();
</script>
</body>
</html>
"""

HTML = (HTML_TEMPLATE
        .replace("__TRIP__", META["trip"])
        .replace("__DATA_JSON__", json.dumps(DATA, ensure_ascii=False))
        .replace("__A__", A)
        .replace("__B__", B))

build_excel(os.path.join(HERE, "trip_account.xlsx"))
with open(os.path.join(HERE, "index.html"), "w", encoding="utf-8") as f:
    f.write(HTML)
print("网页已生成:", os.path.join(HERE, "index.html"))
print("汇总:", json.dumps(AGG, ensure_ascii=False))
